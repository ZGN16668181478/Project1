import openpyxl
from openpyxl.reader.excel import load_workbook

def readExcel(path):
    # 打开文件
    file = load_workbook(filename=path)
    # 打印所有表格的名称
    # print(file.get_sheet_names())
    sheets = file.get_sheet_names()
    # 拿出一个表格
    sheet = file.get_sheet_by_name(sheets[0])
    # 最大行列和表头名
    # print(sheet.max_row)
    # print(sheet.max_column)
    # print(sheet.title)

    # 读取一张表的数据,根据行列数定位拿数据
    # 必须要从第一行开始，这里的最大行列数都是索引值，在数量上需要加一
    for lineNum in range(1,sheet.max_row+1):
        lineList = []
        for columnNum in range(1,sheet.max_column+1):
            value = sheet.cell(row=lineNum,column=columnNum).value
            # if value!=None:
            lineList.append(value)
        print(lineList)

# 利用pyexcel_xls模块的get_data处理两种样式
from pyexcel_xls import get_data
# 有序字典
from collections import OrderedDict

def readAllExcel(path):
    dic = OrderedDict()
    data = get_data(path)
    for sheet in data:
        dic[sheet] = data[sheet]
        # print(dic[sheet])
        print(dic[sheet])
    return dict(dic)
# 下面输出时会有dict这样的标注

path = r'C:\Users\Asus\AppData\Roaming\Kingsoft\wps\addons\data\win-i386\dl\05_nzjgeshui.xlsx'
path2 = r'C:\Users\Asus\Documents\05_nzjgeshui.xls'
# readExcel(path)
res = readAllExcel(path2),
print(res)